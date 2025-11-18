from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.core.exceptions import PermissionDenied
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.core.paginator import Paginator
from django.core.mail import send_mail
from django.template.loader import render_to_string
from django.utils.html import strip_tags
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime
from django.contrib.auth.hashers import make_password
from django.conf import settings
from productos.models import Producto
from inventarios.models import Inventario
from usuarios.models import Usuario, PasswordResetToken
from .forms import ProductoForm, InventarioForm

def login_view(request):
    """Vista de login personalizada"""
    # Si el usuario ya está autenticado, redirigir al dashboard
    if request.user.is_authenticated:
        return redirect('dashboard:home')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        user = authenticate(request, username=username, password=password)
        if user is not None:
            login(request, user)
            
            # Verificar si el usuario debe cambiar su contraseña
            if hasattr(user, 'forzar_cambio_contrasena') and user.forzar_cambio_contrasena:
                # Redirigir a la página de cambio de contraseña obligatorio
                return redirect('dashboard:cambiar_contrasena_obligatorio')
            
            return redirect('dashboard:home')
        else:
            messages.error(request, 'Usuario o contraseña incorrectos')
    
    return render(request, 'dashboard/new_login.html')

def logout_view(request):
    """Vista de logout"""
    logout(request)
    return redirect('dashboard:login')

@login_required
def home(request):
    """Dashboard principal"""
    user = request.user
    now = timezone.now()
    
    # Datos para el contexto
    context = {
        'user': user,
        'productos_count': Producto.objects.count(),
        'inventarios_count': Inventario.objects.count(),
        'today': now.date(),
        'now': now,
    }
    
    # Datos ficticios para proveedores y ventas (solo para administradores)
    if user.is_superuser or (hasattr(user, 'id_rol') and user.id_rol.nombre == 'Administrador'):
        context.update({
            'proveedores_count': 12,  # Ficticio
            'ventas_count': 156,      # Ficticio
        })
    
    return render(request, 'dashboard/home.html', context)

@login_required
def productos_view(request):
    """Vista para listar todos los productos con búsqueda, paginación y ordenamiento"""
    user = request.user
    
    # Verificar permisos según rol
    rol_nombre = user.id_rol.nombre if hasattr(user, 'id_rol') and user.id_rol else None
    es_vendedor = rol_nombre == 'Vendedor'
    es_bodeguero = rol_nombre == 'Bodeguero'
    puede_crear_editar = user.is_superuser or rol_nombre in ['Administrador', 'Bodeguero']
    puede_eliminar = user.is_superuser or rol_nombre == 'Administrador'
    
    productos = Producto.objects.all()
    
    # Búsqueda por múltiples campos
    search = request.GET.get('search', '')
    if search:
        productos = productos.filter(
            nombre__icontains=search
        ) | productos.filter(
            descripcion__icontains=search
        ) | productos.filter(
            precio_referencia__icontains=search
        )
    
    # Ordenamiento
    order_by = request.GET.get('order_by', 'id_producto')
    order_direction = request.GET.get('order_direction', 'asc')
    
    # Construir el campo de ordenamiento
    if order_direction == 'desc':
        order_field = f'-{order_by}' if not order_by.startswith('-') else order_by
    else:
        order_field = order_by.replace('-', '')
    
    productos = productos.order_by(order_field)
    
    # Paginación - obtener de sesión o de parámetro GET
    per_page_param = request.GET.get('per_page')
    if per_page_param:
        per_page = int(per_page_param)
        request.session['productos_per_page'] = per_page
    else:
        per_page = request.session.get('productos_per_page', 10)
        # Asegurar que sea entero
        if isinstance(per_page, str):
            per_page = int(per_page)
    
    paginator = Paginator(productos, per_page)
    page = request.GET.get('page', 1)
    productos_paginados = paginator.get_page(page)
    
    context = {
        'productos': productos_paginados,
        'search': search,
        'order_by': order_by.replace('-', ''),
        'order_direction': order_direction,
        'per_page': per_page,
        'total_productos': Producto.objects.count(),
        'productos_activos': Producto.objects.count(),
        'es_vendedor': es_vendedor,
        'es_bodeguero': es_bodeguero,
        'puede_crear_editar': puede_crear_editar,
        'puede_eliminar': puede_eliminar,
    }
    return render(request, 'dashboard/productos.html', context)

@login_required
def inventarios_view(request):
    """Vista de inventarios"""
    user = request.user
    
    # Verificar permisos según rol
    rol_nombre = user.id_rol.nombre if hasattr(user, 'id_rol') and user.id_rol else None
    es_vendedor = rol_nombre == 'Vendedor'
    es_bodeguero = rol_nombre == 'Bodeguero'
    puede_editar = user.is_superuser or rol_nombre in ['Administrador', 'Bodeguero']
    
    inventarios = Inventario.objects.select_related('id_producto').all()
    now = timezone.now()
    
    # Mock data para proveedores
    class MockProveedor:
        def __init__(self, id_proveedor, nombre):
            self.id_proveedor = id_proveedor
            self.nombre = nombre
    
    proveedores = [
        MockProveedor(1, 'Distribuidora Nacional'),
        MockProveedor(2, 'Dulces Premium SAC'),
        MockProveedor(3, 'Confitería del Norte'),
    ]
    
    context = {
        'inventarios': inventarios,
        'proveedores': proveedores,
        'total_productos': inventarios.count(),
        'stock_alto': 0,  # Calcular basado en lógica de stock
        'stock_medio': 0,
        'stock_bajo': 0,
        'today': now.date(),
        'user': request.user,
        'es_vendedor': es_vendedor,
        'es_bodeguero': es_bodeguero,
        'puede_editar': puede_editar,
    }
    return render(request, 'dashboard/inventarios.html', context)

@login_required
def proveedores_view(request):
    """Vista de gestión de proveedores"""
    user = request.user
    
    # Solo administradores pueden acceder
    if not (user.is_superuser or (hasattr(user, 'id_rol') and user.id_rol.nombre == 'Administrador')):
        raise PermissionDenied("No tienes permisos para acceder a esta sección")
    
    from proveedores.models import Proveedor
    from django.db.models import Q
    
    # Obtener parámetros de búsqueda y filtro
    search = request.GET.get('search', '')
    per_page = request.GET.get('per_page', request.session.get('proveedores_per_page', 10))
    order_by = request.GET.get('order_by', 'id_proveedor')
    order_direction = request.GET.get('order_direction', 'asc')
    
    # Guardar per_page en sesión
    request.session['proveedores_per_page'] = int(per_page)
    
    # Obtener proveedores
    proveedores = Proveedor.objects.all()
    
    # Aplicar búsqueda
    if search:
        proveedores = proveedores.filter(
            Q(nombre__icontains=search) |
            Q(contacto__icontains=search) |
            Q(direccion__icontains=search)
        )
    
    # Aplicar ordenamiento
    order_field = order_by if order_direction == 'asc' else f'-{order_by}'
    proveedores = proveedores.order_by(order_field)
    
    # Aplicar paginación
    paginator = Paginator(proveedores, per_page)
    page_number = request.GET.get('page', 1)
    
    try:
        proveedores_page = paginator.page(page_number)
    except PageNotAnInteger:
        proveedores_page = paginator.page(1)
    except EmptyPage:
        proveedores_page = paginator.page(paginator.num_pages)
    
    # Productos disponibles para asociar con proveedores
    productos_disponibles = Producto.objects.all()
    
    context = {
        'proveedores': proveedores_page,
        'productos_disponibles': productos_disponibles,
        'proveedores_count': paginator.count,
        'proveedores_activos': paginator.count,
        'productos_proveedor': productos_disponibles.count(),
        'ordenes_pendientes': 0,
        'search': search,
        'per_page': int(per_page),
        'order_by': order_by,
        'order_direction': order_direction,
        'user': request.user,
    }
    return render(request, 'dashboard/proveedores.html', context)

@login_required
def obtener_proveedor(request, proveedor_id):
    """API para obtener datos de un proveedor en formato JSON"""
    user = request.user
    
    # Solo administradores pueden acceder
    if not (user.is_superuser or (hasattr(user, 'id_rol') and user.id_rol.nombre == 'Administrador')):
        return JsonResponse({'success': False, 'message': 'No tienes permisos'}, status=403)
    
    try:
        from proveedores.models import Proveedor
        from producto_proveedor.models import ProductoProveedor
        
        proveedor = Proveedor.objects.get(id_proveedor=proveedor_id)
        
        # Obtener productos asociados con sus nombres
        productos_relaciones = ProductoProveedor.objects.filter(id_proveedor=proveedor).select_related('id_producto')
        productos_ids = [rel.id_producto.id_producto for rel in productos_relaciones]
        productos_nombres = [{'id': rel.id_producto.id_producto, 'nombre': rel.id_producto.nombre} for rel in productos_relaciones]
        
        # Parsear dirección para obtener comuna y región
        direccion_completa = proveedor.direccion or ''
        partes_direccion = direccion_completa.split(', ')
        direccion_calle = partes_direccion[0] if len(partes_direccion) > 0 else ''
        comuna = partes_direccion[1] if len(partes_direccion) > 1 else ''
        region = partes_direccion[2] if len(partes_direccion) > 2 else ''
        
        data = {
            'success': True,
            'proveedor': {
                'id': proveedor.id_proveedor,
                'nombre': proveedor.nombre,
                'contacto': proveedor.contacto,
                'direccion': direccion_calle,
                'direccion_completa': direccion_completa,
                'comuna': comuna,
                'region': region,
                'productos': productos_ids,
                'productos_detalle': productos_nombres
            }
        }
        return JsonResponse(data)
    except Proveedor.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Proveedor no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'}, status=500)

@login_required
def guardar_proveedor(request):
    """API para crear o actualizar un proveedor"""
    user = request.user
    
    # Solo administradores pueden acceder
    if not (user.is_superuser or (hasattr(user, 'id_rol') and user.id_rol.nombre == 'Administrador')):
        return JsonResponse({'success': False, 'message': 'No tienes permisos'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Método no permitido'}, status=405)
    
    try:
        from proveedores.models import Proveedor
        from producto_proveedor.models import ProductoProveedor
        import json
        
        proveedor_id = request.POST.get('proveedor_id')
        nombre = request.POST.get('nombre', '').strip()
        contacto = request.POST.get('contacto', '').strip()
        direccion = request.POST.get('direccion', '').strip()
        comuna = request.POST.get('comuna', '').strip()
        region = request.POST.get('region', '').strip()
        tipo_proveedor = request.POST.get('tipo_proveedor', '').strip()
        condiciones_pago = request.POST.get('condiciones_pago', '').strip()
        tiempo_entrega = request.POST.get('tiempo_entrega', '').strip()
        monto_minimo = request.POST.get('monto_minimo', '').strip()
        productos_ids = request.POST.getlist('productos[]')
        
        # Validaciones
        if not nombre:
            return JsonResponse({'success': False, 'message': 'El nombre es requerido'}, status=400)
        
        if len(nombre) > 150:
            return JsonResponse({'success': False, 'message': 'El nombre no puede exceder 150 caracteres'}, status=400)
        
        if len(contacto) > 200:
            return JsonResponse({'success': False, 'message': 'El contacto no puede exceder 200 caracteres'}, status=400)
        
        # Construir dirección completa con información chilena
        direccion_completa = direccion
        if comuna:
            direccion_completa += f", {comuna}"
        if region:
            direccion_completa += f", {region}"
        
        # Guardar datos adicionales en JSON para recuperarlos después
        datos_adicionales = {
            'direccion_calle': direccion,
            'comuna': comuna,
            'region': region,
            'tipo_proveedor': tipo_proveedor,
            'condiciones_pago': condiciones_pago,
            'tiempo_entrega': tiempo_entrega,
            'monto_minimo': monto_minimo
        }
        
        # Crear o actualizar proveedor
        if proveedor_id:
            # Actualizar existente
            proveedor = Proveedor.objects.get(id_proveedor=proveedor_id)
            proveedor.nombre = nombre
            proveedor.contacto = contacto
            proveedor.direccion = direccion_completa[:200]  # Limitar a 200 caracteres
            proveedor.save()
            mensaje = f'Proveedor "{nombre}" actualizado exitosamente'
            
            # Eliminar asociaciones anteriores
            ProductoProveedor.objects.filter(id_proveedor=proveedor).delete()
        else:
            # Crear nuevo
            proveedor = Proveedor.objects.create(
                nombre=nombre,
                contacto=contacto,
                direccion=direccion_completa[:200]
            )
            mensaje = f'Proveedor "{nombre}" creado exitosamente'
        
        # Asociar productos seleccionados
        productos_asociados = 0
        if productos_ids:
            for producto_id in productos_ids:
                try:
                    producto = Producto.objects.get(id_producto=producto_id)
                    ProductoProveedor.objects.create(
                        id_producto=producto,
                        id_proveedor=proveedor,
                        precio_acordado=0  # Valor por defecto, puede editarse después
                    )
                    productos_asociados += 1
                except Producto.DoesNotExist:
                    continue
        
        if productos_asociados > 0:
            mensaje += f' con {productos_asociados} producto(s) asociado(s)'
        
        return JsonResponse({
            'success': True,
            'message': mensaje,
            'proveedor': {
                'id': proveedor.id_proveedor,
                'nombre': proveedor.nombre,
                'contacto': proveedor.contacto,
                'direccion': proveedor.direccion,
                'datos_adicionales': datos_adicionales,
                'productos_asociados': productos_asociados
            }
        })
        
    except Proveedor.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Proveedor no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error al guardar: {str(e)}'}, status=500)

@login_required
def eliminar_proveedor(request, proveedor_id):
    """API para eliminar un proveedor"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Método no permitido'}, status=405)
    
    user = request.user
    
    # Verificar permisos - solo administradores
    if not (user.is_superuser or (hasattr(user, 'id_rol') and user.id_rol.nombre == 'Administrador')):
        return JsonResponse({'success': False, 'message': 'No tienes permisos para eliminar proveedores'}, status=403)
    
    try:
        from proveedores.models import Proveedor
        
        # Obtener proveedor
        proveedor = Proveedor.objects.get(id_proveedor=proveedor_id)
        nombre_proveedor = proveedor.nombre
        
        # Eliminar proveedor
        proveedor.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Proveedor "{nombre_proveedor}" eliminado exitosamente'
        })
        
    except Proveedor.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Proveedor no encontrado'}, status=404)
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error al eliminar: {str(e)}'}, status=500)

@login_required
def exportar_proveedores_excel(request):
    """Exportar proveedores a Excel con formato profesional"""
    user = request.user
    
    # Solo administradores pueden exportar
    if not (user.is_superuser or (hasattr(user, 'id_rol') and user.id_rol.nombre == 'Administrador')):
        return JsonResponse({'success': False, 'message': 'No tienes permisos'}, status=403)
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse
        from datetime import datetime
        from proveedores.models import Proveedor
        from producto_proveedor.models import ProductoProveedor
        
        # Crear workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Proveedores"
        
        # Estilos
        header_fill = PatternFill(start_color="4F81F7", end_color="4F81F7", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Título
        ws.merge_cells('A1:F1')
        ws['A1'] = '🏢 LISTADO DE PROVEEDORES - DULCERÍA LILIS'
        ws['A1'].font = Font(bold=True, size=16, color="4F81F7")
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # Fecha de exportación
        ws.merge_cells('A2:F2')
        ws['A2'] = f'Exportado el: {datetime.now().strftime("%d/%m/%Y %H:%M:%S")}'
        ws['A2'].alignment = Alignment(horizontal='center')
        ws['A2'].font = Font(italic=True, size=10)
        ws.row_dimensions[2].height = 20
        
        # Encabezados
        headers = ['ID', 'Nombre', 'Contacto', 'Dirección', 'Comuna', 'Región', 'Productos']
        ws.append([])
        ws.append(headers)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=4, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        ws.row_dimensions[4].height = 25
        
        # Datos
        proveedores = Proveedor.objects.all().order_by('id_proveedor')
        
        for proveedor in proveedores:
            # Parsear dirección
            partes_direccion = proveedor.direccion.split(', ') if proveedor.direccion else []
            direccion_calle = partes_direccion[0] if len(partes_direccion) > 0 else ''
            comuna = partes_direccion[1] if len(partes_direccion) > 1 else ''
            region = partes_direccion[2] if len(partes_direccion) > 2 else ''
            
            # Obtener productos asociados
            productos = ProductoProveedor.objects.filter(id_proveedor=proveedor).select_related('id_producto')
            productos_nombres = ', '.join([p.id_producto.nombre for p in productos]) if productos.exists() else 'Sin productos'
            
            row = [
                proveedor.id_proveedor,
                proveedor.nombre,
                proveedor.contacto or 'Sin contacto',
                direccion_calle or 'Sin dirección',
                comuna or '-',
                region or '-',
                productos_nombres
            ]
            ws.append(row)
            
            # Aplicar bordes y alineación a cada celda
            current_row = ws.max_row
            for col_num in range(1, len(headers) + 1):
                cell = ws.cell(row=current_row, column=col_num)
                cell.border = border
                cell.alignment = Alignment(vertical='center', wrap_text=True)
                
                # ID centrado
                if col_num == 1:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Ajustar anchos de columna
        column_widths = [8, 30, 25, 35, 20, 30, 40]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
        
        # Agregar totales
        total_row = ws.max_row + 2
        ws.merge_cells(f'A{total_row}:C{total_row}')
        ws[f'A{total_row}'] = f'Total de Proveedores: {proveedores.count()}'
        ws[f'A{total_row}'].font = Font(bold=True, size=11)
        ws[f'A{total_row}'].alignment = Alignment(horizontal='left')
        
        # Preparar respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="Proveedores_Lilis_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error al exportar: {str(e)}'}, status=500)

@login_required
def ventas_view(request):
    """Vista ficticia de ventas"""
    user = request.user
    
    # Solo administradores pueden acceder
    if not (user.is_superuser or (hasattr(user, 'id_rol') and user.id_rol.nombre == 'Administrador')):
        raise PermissionDenied("No tienes permisos para acceder a esta sección")
    
    # Datos ficticios
    ventas = [
        {'id': 1, 'fecha': '2025-01-15', 'cliente': 'Cliente A', 'total': 50000, 'estado': 'Completada'},
        {'id': 2, 'fecha': '2025-01-14', 'cliente': 'Cliente B', 'total': 75000, 'estado': 'Pendiente'},
        {'id': 3, 'fecha': '2025-01-13', 'cliente': 'Cliente C', 'total': 30000, 'estado': 'Completada'},
    ]
    
    context = {
        'ventas': ventas,
        'user': request.user,
    }
    return render(request, 'dashboard/ventas.html', context)

@login_required
def agregar_producto(request):
    """Vista para agregar un nuevo producto"""
    user = request.user
    
    # Verificar permisos
    if not (user.is_superuser or (hasattr(user, 'id_rol') and user.id_rol.nombre in ['Administrador', 'Bodeguero'])):
        messages.error(request, 'No tienes permisos para agregar productos')
        return redirect('dashboard:productos')
    
    if request.method == 'POST':
        form = ProductoForm(request.POST)
        if form.is_valid():
            producto = form.save()
            messages.success(request, f'Producto "{producto.nombre}" agregado exitosamente')
            return redirect('dashboard:productos')
        else:
            messages.error(request, 'Por favor corrige los errores en el formulario')
    else:
        form = ProductoForm()
    
    context = {
        'form': form,
        'user': user,
        'titulo': 'Agregar Nuevo Producto'
    }
    return render(request, 'dashboard/form_producto.html', context)

@login_required
def editar_producto(request, producto_id):
    """Vista para editar un producto existente"""
    user = request.user
    producto = get_object_or_404(Producto, id_producto=producto_id)
    
    # Solo administradores pueden editar
    if not (user.is_superuser or (hasattr(user, 'id_rol') and user.id_rol.nombre == 'Administrador')):
        messages.error(request, 'No tienes permisos para editar productos')
        return redirect('dashboard:productos')
    
    if request.method == 'POST':
        form = ProductoForm(request.POST, instance=producto)
        if form.is_valid():
            form.save()
            messages.success(request, 'Producto actualizado exitosamente')
            return redirect('dashboard:productos')
    else:
        form = ProductoForm(instance=producto)
    
    context = {
        'form': form,
        'user': user,
        'titulo': f'Editar Producto: {producto.nombre}',
        'producto': producto
    }
    return render(request, 'dashboard/form_producto.html', context)

@login_required
def agregar_inventario(request):
    """Vista para agregar un nuevo inventario"""
    user = request.user
    
    # Solo administradores pueden agregar inventarios
    if not (user.is_superuser or (hasattr(user, 'id_rol') and user.id_rol.nombre == 'Administrador')):
        messages.error(request, 'Solo los administradores pueden agregar inventarios')
        return redirect('dashboard:inventarios')
    
    if request.method == 'POST':
        form = InventarioForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Inventario agregado exitosamente')
            return redirect('dashboard:inventarios')
    else:
        form = InventarioForm()
    
    context = {
        'form': form,
        'user': user,
        'titulo': 'Agregar Nuevo Inventario'
    }
    return render(request, 'dashboard/form_inventario.html', context)

@login_required
def editar_inventario(request, inventario_id):
    """Vista para editar un inventario existente"""
    user = request.user
    inventario = get_object_or_404(Inventario, id_inventario=inventario_id)
    
    # Solo administradores pueden editar
    if not (user.is_superuser or (hasattr(user, 'id_rol') and user.id_rol.nombre == 'Administrador')):
        messages.error(request, 'No tienes permisos para editar inventarios')
        return redirect('dashboard:inventarios')
    
    if request.method == 'POST':
        form = InventarioForm(request.POST, instance=inventario)
        if form.is_valid():
            form.save()
            messages.success(request, 'Inventario actualizado exitosamente')
            return redirect('dashboard:inventarios')
    else:
        form = InventarioForm(instance=inventario)
    
    context = {
        'form': form,
        'user': user,
        'titulo': f'Editar Inventario: {inventario.id_producto.nombre}',
        'inventario': inventario
    }
    return render(request, 'dashboard/form_inventario.html', context)

def forgot_password_view(request):
    """Vista para recuperación de contraseña"""
    if request.method == 'POST':
        email = request.POST.get('email')
        
        try:
            usuario = Usuario.objects.get(correo=email)
            
            # Crear token de recuperación
            token = PasswordResetToken.objects.create(usuario=usuario)
            
            # Construir URL de recuperación
            reset_url = request.build_absolute_uri(
                f'/reset-password/?token={token.token}'
            )
            
            # Renderizar template de email
            html_message = render_to_string('dashboard/password_reset_email.html', {
                'usuario': usuario,
                'reset_url': reset_url,
                'token': token
            })
            plain_message = strip_tags(html_message)
            
            # Enviar email
            send_mail(
                subject='Recuperación de Contraseña - Dulcería Lilis',
                message=plain_message,
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[email],
                html_message=html_message,
                fail_silently=False,
            )
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': 'Se ha enviado un email con las instrucciones para recuperar tu contraseña'
                })
            
            messages.success(request, 'Se ha enviado un email con las instrucciones para recuperar tu contraseña')
            return redirect('dashboard:login')
            
        except Usuario.DoesNotExist:
            # Por seguridad, no revelamos que el email no existe
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': 'Si el email existe, recibirás instrucciones para recuperar tu contraseña'
                })
            
            messages.info(request, 'Si el email existe, recibirás instrucciones para recuperar tu contraseña')
            return redirect('dashboard:login')
            
        except Exception as e:
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'message': 'Error al enviar el email. Por favor intenta nuevamente.'
                })
            
            messages.error(request, 'Error al enviar el email. Por favor intenta nuevamente.')
    
    return render(request, 'dashboard/forgot_password.html')

def reset_password_view(request):
    """Vista para resetear contraseña con token"""
    token_str = request.GET.get('token') or request.POST.get('token')
    
    if not token_str:
        messages.error(request, 'Token de recuperación no válido')
        return redirect('dashboard:login')
    
    try:
        token = PasswordResetToken.objects.get(token=token_str)
        
        if not token.is_valid():
            messages.error(request, 'El token ha expirado o ya fue usado. Solicita uno nuevo.')
            return redirect('dashboard:forgot_password')
        
        if request.method == 'POST':
            password = request.POST.get('password')
            password_confirm = request.POST.get('password_confirm')
            
            if password != password_confirm:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': False,
                        'message': 'Las contraseñas no coinciden'
                    })
                messages.error(request, 'Las contraseñas no coinciden')
            elif len(password) < 8:
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': False,
                        'message': 'La contraseña debe tener al menos 8 caracteres'
                    })
                messages.error(request, 'La contraseña debe tener al menos 8 caracteres')
            else:
                # Cambiar contraseña
                usuario = token.usuario
                usuario.password = make_password(password)
                usuario.save()
                
                # Marcar token como usado
                token.is_used = True
                token.save()
                
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': True,
                        'message': 'Contraseña cambiada exitosamente'
                    })
                
                messages.success(request, 'Contraseña cambiada exitosamente. Ahora puedes iniciar sesión.')
                return redirect('dashboard:login')
        
        context = {
            'token': token_str,
            'usuario': token.usuario
        }
        return render(request, 'dashboard/reset_password.html', context)
        
    except PasswordResetToken.DoesNotExist:
        messages.error(request, 'Token de recuperación no válido')
        return redirect('dashboard:login')


@login_required
def usuarios_view(request):
    """Vista de gestión de usuarios con búsqueda, paginación y ordenamiento"""
    user = request.user
    
    # Solo administradores pueden gestionar usuarios
    if not (user.is_superuser or (hasattr(user, 'id_rol') and user.id_rol.nombre == 'Administrador')):
        raise PermissionDenied("No tienes permisos para gestionar usuarios")
    
    # Usar el modelo de Usuario personalizado
    from usuarios.models import Usuario
    from roles.models import Rol
    usuarios = Usuario.objects.select_related('id_rol').all()
    roles = Rol.objects.all()
    
    # Búsqueda por múltiples campos
    search = request.GET.get('search', '')
    if search:
        usuarios = usuarios.filter(
            nombre__icontains=search
        ) | usuarios.filter(
            username__icontains=search
        ) | usuarios.filter(
            email__icontains=search
        ) | usuarios.filter(
            correo__icontains=search
        )
    
    # Ordenamiento
    order_by = request.GET.get('order_by', 'id_usuario')
    order_direction = request.GET.get('order_direction', 'asc')
    
    # Construir el campo de ordenamiento
    if order_direction == 'desc':
        order_field = f'-{order_by}' if not order_by.startswith('-') else order_by
    else:
        order_field = order_by.replace('-', '')
    
    usuarios = usuarios.order_by(order_field)
    
    # Paginación - obtener de sesión o de parámetro GET
    per_page_param = request.GET.get('per_page')
    if per_page_param:
        per_page = int(per_page_param)
        request.session['usuarios_per_page'] = per_page
    else:
        per_page = request.session.get('usuarios_per_page', 10)
        # Asegurar que sea entero
        if isinstance(per_page, str):
            per_page = int(per_page)
    
    paginator = Paginator(usuarios, per_page)
    page = request.GET.get('page', 1)
    usuarios_paginados = paginator.get_page(page)
    
    context = {
        'usuarios': usuarios_paginados,
        'roles': roles,
        'search': search,
        'order_by': order_by.replace('-', ''),
        'order_direction': order_direction,
        'per_page': per_page,
        'usuarios_count': Usuario.objects.count(),
        'usuarios_activos': Usuario.objects.filter(is_active=True).count(),
        'usuarios_inactivos': Usuario.objects.filter(is_active=False).count(),
        'nuevos_usuarios': 0,  # Mock data
        'user': request.user,
    }
    return render(request, 'dashboard/usuarios.html', context)

@login_required
def obtener_usuario(request, usuario_id):
    """API para obtener datos de un usuario en formato JSON"""
    import traceback
    
    # Verificar autenticación
    if not request.user.is_authenticated:
        return JsonResponse({'success': False, 'message': 'No autenticado'}, status=401)
    
    user = request.user
    
    # Solo administradores pueden acceder
    try:
        if not (user.is_superuser or (hasattr(user, 'id_rol') and user.id_rol.nombre == 'Administrador')):
            return JsonResponse({'success': False, 'message': 'No tienes permisos'}, status=403)
    except Exception as e:
        return JsonResponse({'success': False, 'message': f'Error verificando permisos: {str(e)}'}, status=500)
    
    try:
        from usuarios.models import Usuario
        usuario = Usuario.objects.select_related('id_rol').get(id_usuario=usuario_id)
        
        data = {
            'success': True,
            'usuario': {
                'id': usuario.id_usuario,
                'usuario': usuario.username,
                'email': usuario.email,
                'nombre': usuario.nombre,
                'telefono': usuario.telefono or '',
                'id_rol': usuario.id_rol.id_rol if usuario.id_rol else '',
                'is_active': usuario.is_active,
                'cambiar_password': False  # Por defecto no forzar cambio
            }
        }
        return JsonResponse(data)
    except Usuario.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Usuario no encontrado'}, status=404)
    except Exception as e:
        error_trace = traceback.format_exc()
        print(f"Error en obtener_usuario: {error_trace}")
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'}, status=500)

@login_required
def guardar_usuario(request):
    """API para crear o actualizar un usuario"""
    user = request.user
    
    # Solo administradores pueden acceder
    if not (user.is_superuser or (hasattr(user, 'id_rol') and user.id_rol.nombre == 'Administrador')):
        return JsonResponse({'success': False, 'message': 'No tienes permisos'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Método no permitido'}, status=405)
    
    try:
        from usuarios.models import Usuario
        from roles.models import Rol
        import secrets
        import string
        
        usuario_id = request.POST.get('user_id')
        email = request.POST.get('email', '').strip()
        nombre = request.POST.get('nombre', '').strip()
        telefono = request.POST.get('telefono', '').strip()
        id_rol = request.POST.get('id_rol')
        # El checkbox puede venir como 'on', 'true', o no venir
        activo = request.POST.get('activo', 'off') in ['on', 'true', True]
        forzar_cambio = request.POST.get('forzar_cambio_contrasena', 'on') in ['on', 'true', True]
        
        # Validaciones básicas
        print(f"DEBUG - Datos recibidos: email={email}, nombre={nombre}, id_rol={id_rol}")
        
        if not email:
            return JsonResponse({
                'success': False, 
                'errors': {'email': ['El campo Email es requerido']}
            })
        
        if not nombre:
            return JsonResponse({
                'success': False, 
                'errors': {'nombre': ['El campo Nombre es requerido']}
            })
        
        if not id_rol:
            return JsonResponse({
                'success': False, 
                'errors': {'id_rol': ['El campo Rol es requerido']}
            })
        
        # Obtener el rol
        try:
            rol = Rol.objects.get(pk=id_rol)
        except Rol.DoesNotExist:
            return JsonResponse({
                'success': False,
                'errors': {'id_rol': ['El rol seleccionado no existe']}
            })
        
        # Modo edición o creación
        if usuario_id:
            # Editar usuario existente
            usuario = get_object_or_404(Usuario, id_usuario=usuario_id)
            
            # Verificar que el email no esté siendo usado por otro usuario
            if Usuario.objects.filter(correo=email).exclude(id_usuario=usuario_id).exists():
                return JsonResponse({
                    'success': False,
                    'errors': {'email': ['Este correo electrónico ya está registrado']}
                })
            
            # Actualizar datos
            usuario.username = email
            usuario.email = email
            usuario.correo = email
            usuario.nombre = nombre
            usuario.telefono = telefono if telefono else None
            usuario.id_rol = rol
            usuario.is_active = activo
            
            usuario.save()
            action = 'actualizado'
            temp_password = None  # No se genera contraseña en edición
        else:
            # Crear nuevo usuario
            # Verificar que el email no exista
            if Usuario.objects.filter(correo=email).exists():
                return JsonResponse({
                    'success': False,
                    'errors': {'email': ['Este correo electrónico ya está registrado']}
                })
            
            # Generar contraseña temporal aleatoria (12 caracteres con letras, números y símbolos)
            alphabet = string.ascii_letters + string.digits + "!@#$%&*"
            temp_password = ''.join(secrets.choice(alphabet) for i in range(12))
            
            # Crear usuario
            usuario = Usuario(
                username=email,
                email=email,
                correo=email,
                nombre=nombre,
                telefono=telefono if telefono else None,
                id_rol=rol,
                is_active=activo,
                forzar_cambio_contrasena=forzar_cambio
            )
            usuario.set_password(temp_password)
            usuario.save()
            action = 'creado'
        
        # Preparar respuesta
        response_data = {
            'success': True,
            'message': f'Usuario {action} correctamente',
            'usuario': {
                'id': usuario.id_usuario,
                'usuario': usuario.username,
                'nombre': usuario.nombre,
                'email': usuario.email
            }
        }
        
        # Si es un nuevo usuario, incluir la contraseña temporal
        if temp_password:
            response_data['temp_password'] = temp_password
            response_data['message'] = f'Usuario creado correctamente. Contraseña temporal: {temp_password}'
        
        return JsonResponse(response_data)
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'errors': {'general': [str(e)]}
        }, status=500)

@login_required
def eliminar_usuario(request, usuario_id):
    """API para eliminar un usuario"""
    user = request.user
    
    # Solo administradores pueden acceder
    if not (user.is_superuser or (hasattr(user, 'id_rol') and user.id_rol.nombre == 'Administrador')):
        return JsonResponse({'success': False, 'message': 'No tienes permisos para realizar esta acción'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Método no permitido'}, status=405)
    
    try:
        from usuarios.models import Usuario
        usuario = get_object_or_404(Usuario, id_usuario=usuario_id)
        
        # Verificar si el usuario puede ser eliminado
        if not usuario.can_be_deleted():
            return JsonResponse({
                'success': False,
                'message': 'No se puede eliminar este usuario porque es un administrador del sistema'
            })
        
        # No permitir eliminar el usuario actual
        if usuario.id_usuario == user.id_usuario:
            return JsonResponse({
                'success': False,
                'message': 'No puedes eliminar tu propia cuenta'
            })
        
        nombre = usuario.nombre
        usuario.delete()
        
        return JsonResponse({
            'success': True,
            'message': f'Usuario "{nombre}" eliminado correctamente'
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error al eliminar el usuario: {str(e)}'
        }, status=500)

@login_required
def cambiar_contrasena_obligatorio(request):
    """Vista para cambio de contraseña obligatorio en el primer login"""
    user = request.user
    
    # Si el usuario no necesita cambiar la contraseña, redirigir al dashboard
    if not hasattr(user, 'forzar_cambio_contrasena') or not user.forzar_cambio_contrasena:
        return redirect('dashboard:home')
    
    if request.method == 'POST':
        nueva_contrasena = request.POST.get('nueva_contrasena', '')
        confirmar_contrasena = request.POST.get('confirmar_contrasena', '')
        
        # Validaciones
        if not nueva_contrasena:
            messages.error(request, 'La nueva contraseña es requerida')
        elif len(nueva_contrasena) < 8:
            messages.error(request, 'La contraseña debe tener al menos 8 caracteres')
        elif nueva_contrasena != confirmar_contrasena:
            messages.error(request, 'Las contraseñas no coinciden')
        else:
            # Cambiar la contraseña
            user.set_password(nueva_contrasena)
            user.forzar_cambio_contrasena = False
            user.save()
            
            # Re-autenticar al usuario con la nueva contraseña
            from django.contrib.auth import update_session_auth_hash
            update_session_auth_hash(request, user)
            
            messages.success(request, 'Contraseña cambiada exitosamente')
            return redirect('dashboard:home')
    
    return render(request, 'dashboard/cambiar_contrasena_obligatorio.html', {
        'user': user
    })

@login_required
def cambiar_estado_usuario(request, usuario_id):
    """API para activar/desactivar un usuario"""
    user = request.user
    
    # Solo administradores pueden acceder
    if not (user.is_superuser or (hasattr(user, 'id_rol') and user.id_rol.nombre == 'Administrador')):
        return JsonResponse({'success': False, 'message': 'No tienes permisos para realizar esta acción'}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Método no permitido'}, status=405)
    
    try:
        from usuarios.models import Usuario
        usuario = get_object_or_404(Usuario, id_usuario=usuario_id)
        
        # No permitir cambiar el estado del usuario actual
        if usuario.id_usuario == user.id_usuario:
            return JsonResponse({
                'success': False,
                'message': 'No puedes cambiar el estado de tu propia cuenta'
            })
        
        # Obtener el nuevo estado
        nuevo_estado = request.POST.get('activo', 'false') == 'true'
        
        # Si se intenta desactivar, verificar si puede ser desactivado
        if not nuevo_estado and not usuario.can_be_deactivated():
            return JsonResponse({
                'success': False,
                'message': 'No se puede desactivar este usuario porque es un administrador del sistema'
            })
        
        # Cambiar estado
        usuario.is_active = nuevo_estado
        usuario.save()
        
        estado_texto = 'activado' if nuevo_estado else 'desactivado'
        
        return JsonResponse({
            'success': True,
            'message': f'Usuario "{usuario.nombre}" {estado_texto} correctamente',
            'nuevo_estado': nuevo_estado
        })
        
    except Exception as e:
        return JsonResponse({
            'success': False,
            'message': f'Error al cambiar el estado: {str(e)}'
        }, status=500)

@login_required
def exportar_usuarios_excel(request):
    """Exportar lista de usuarios a Excel"""
    user = request.user
    
    # Solo administradores pueden acceder
    if not (user.is_superuser or (hasattr(user, 'id_rol') and user.id_rol.nombre == 'Administrador')):
        return JsonResponse({'success': False, 'message': 'No tienes permisos'}, status=403)
    
    try:
        from usuarios.models import Usuario
        
        # Crear libro de Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Usuarios"
        
        # Estilos
        header_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Encabezados
        headers = ['ID', 'Usuario', 'Email', 'Nombre', 'Teléfono', 'Rol', 'Estado', 'Último Acceso', 'Fecha Creación']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Obtener usuarios
        usuarios = Usuario.objects.select_related('id_rol').all().order_by('-date_joined')
        
        # Llenar datos
        for row_num, usuario in enumerate(usuarios, 2):
            ws.cell(row=row_num, column=1).value = usuario.id_usuario
            ws.cell(row=row_num, column=2).value = usuario.username
            ws.cell(row=row_num, column=3).value = usuario.email
            ws.cell(row=row_num, column=4).value = usuario.nombre
            ws.cell(row=row_num, column=5).value = usuario.telefono or ''
            ws.cell(row=row_num, column=6).value = usuario.id_rol.nombre if usuario.id_rol else 'Sin rol'
            ws.cell(row=row_num, column=7).value = 'Activo' if usuario.is_active else 'Inactivo'
            ws.cell(row=row_num, column=8).value = usuario.last_login.strftime('%Y-%m-%d %H:%M') if usuario.last_login else 'Nunca'
            ws.cell(row=row_num, column=9).value = usuario.date_joined.strftime('%Y-%m-%d %H:%M') if usuario.date_joined else ''
            
            # Aplicar bordes
            for col_num in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col_num).border = border
                ws.cell(row=row_num, column=col_num).alignment = Alignment(vertical='center')
        
        # Ajustar ancho de columnas
        column_widths = [8, 20, 30, 25, 15, 20, 12, 20, 20]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width
        
        # Crear respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        fecha_actual = datetime.now().strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = f'attachment; filename=usuarios_{fecha_actual}.xlsx'
        
        # Guardar y devolver
        wb.save(response)
        return response
        
    except Exception as e:
        import traceback
        print(f"Error exportando usuarios: {traceback.format_exc()}")
        return JsonResponse({
            'success': False,
            'message': f'Error al exportar: {str(e)}'
        }, status=500)

@login_required
def exportar_productos_excel(request):
    """Exportar lista de productos a Excel"""
    user = request.user
    
    # Solo administradores pueden acceder
    if not (user.is_superuser or (hasattr(user, 'id_rol') and user.id_rol.nombre == 'Administrador')):
        return JsonResponse({'success': False, 'message': 'No tienes permisos'}, status=403)
    
    try:
        from productos.models import Producto
        
        # Crear libro de Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Productos"
        
        # Estilos
        header_fill = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Encabezados
        headers = ['ID', 'Nombre', 'Descripción', 'Precio Referencia']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Obtener productos
        productos = Producto.objects.all().order_by('nombre')
        
        # Llenar datos
        for row_num, producto in enumerate(productos, 2):
            ws.cell(row=row_num, column=1).value = producto.id_producto
            ws.cell(row=row_num, column=2).value = producto.nombre
            ws.cell(row=row_num, column=3).value = producto.descripcion
            ws.cell(row=row_num, column=4).value = producto.precio_referencia
            
            # Formatear precio como moneda
            ws.cell(row=row_num, column=4).number_format = '$#,##0'
            
            # Aplicar bordes
            for col_num in range(1, len(headers) + 1):
                ws.cell(row=row_num, column=col_num).border = border
                ws.cell(row=row_num, column=col_num).alignment = Alignment(vertical='center')
        
        # Ajustar ancho de columnas
        column_widths = [8, 35, 50, 18]
        for col_num, width in enumerate(column_widths, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_num)].width = width
        
        # Crear respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        fecha_actual = datetime.now().strftime('%Y%m%d_%H%M%S')
        response['Content-Disposition'] = f'attachment; filename=productos_{fecha_actual}.xlsx'
        
        # Guardar y devolver
        wb.save(response)
        return response
        
    except Exception as e:
        import traceback
        print(f"Error exportando productos: {traceback.format_exc()}")
        return JsonResponse({
            'success': False,
            'message': f'Error al exportar: {str(e)}'
        }, status=500)

@login_required
def obtener_producto(request, producto_id):
    """API para obtener detalles de un producto en formato JSON"""
    import traceback
    
    # Verificar autenticación
    if not request.user.is_authenticated:
        return JsonResponse({'success': False, 'message': 'No autenticado'}, status=401)
    
    try:
        from productos.models import Producto
        from producto_proveedor.models import ProductoProveedor
        
        producto = Producto.objects.get(id_producto=producto_id)
        
        # Obtener proveedores asociados
        proveedores_asociados = ProductoProveedor.objects.filter(
            id_producto=producto
        ).select_related('id_proveedor')
        
        proveedores_data = []
        for pp in proveedores_asociados:
            proveedores_data.append({
                'id': pp.id_proveedor.id_proveedor,
                'nombre': pp.id_proveedor.nombre,
                'contacto': pp.id_proveedor.contacto if hasattr(pp.id_proveedor, 'contacto') else '',
                'precio_acordado': pp.precio_acordado,
                'fecha_registro': pp.fecha_registro.strftime('%Y-%m-%d') if pp.fecha_registro else ''
            })
        
        data = {
            'success': True,
            'producto': {
                'id': producto.id_producto,
                'nombre': producto.nombre,
                'descripcion': producto.descripcion,
                'precio_referencia': producto.precio_referencia,
                'proveedores': proveedores_data
            }
        }
        return JsonResponse(data)
    except Producto.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Producto no encontrado'}, status=404)
    except Exception as e:
        error_trace = traceback.format_exc()
        print(f"Error en obtener_producto: {error_trace}")
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'}, status=500)

@login_required
def actualizar_producto(request):
    """API para actualizar un producto"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'message': 'Método no permitido'}, status=405)
    
    # Verificar permisos: Administrador y Bodeguero pueden editar
    user = request.user
    rol_nombre = user.id_rol.nombre if hasattr(user, 'id_rol') and user.id_rol else None
    
    if rol_nombre not in ['Administrador', 'Bodeguero'] and not user.is_superuser:
        return JsonResponse({
            'success': False, 
            'message': 'No tienes permisos para editar productos.'
        }, status=403)
    
    try:
        from productos.models import Producto
        
        producto_id = request.POST.get('product_id')
        nombre = request.POST.get('nombre', '').strip()
        descripcion = request.POST.get('descripcion', '').strip()
        precio_referencia = request.POST.get('precio_referencia')
        
        # Validaciones
        if not producto_id:
            return JsonResponse({'success': False, 'message': 'ID de producto no proporcionado'})
        
        if not nombre:
            return JsonResponse({'success': False, 'message': 'El nombre es requerido'})
        
        if not precio_referencia or int(precio_referencia) <= 0:
            return JsonResponse({'success': False, 'message': 'El precio debe ser mayor a 0'})
        
        # Actualizar producto
        producto = Producto.objects.get(id_producto=producto_id)
        producto.nombre = nombre
        producto.descripcion = descripcion
        producto.precio_referencia = int(precio_referencia)
        producto.save()
        
        return JsonResponse({
            'success': True,
            'message': 'Producto actualizado correctamente',
            'producto': {
                'id': producto.id_producto,
                'nombre': producto.nombre,
                'descripcion': producto.descripcion,
                'precio_referencia': producto.precio_referencia
            }
        })
        
    except Producto.DoesNotExist:
        return JsonResponse({'success': False, 'message': 'Producto no encontrado'}, status=404)
    except Exception as e:
        import traceback
        print(f"Error actualizando producto: {traceback.format_exc()}")
        return JsonResponse({'success': False, 'message': f'Error: {str(e)}'}, status=500)


# ==========================================
# Vistas de Error Personalizadas
# ==========================================

def error_404(request, exception=None):
    """Vista personalizada para error 404 - Página no encontrada"""
    return render(request, 'dashboard/error_404.html', status=404)


def error_500(request):
    """Vista personalizada para error 500 - Error del servidor"""
    return render(request, 'dashboard/error_500.html', status=500)


def error_403(request, exception=None):
    """Vista personalizada para error 403 - Acceso denegado"""
    return render(request, 'dashboard/error_403.html', status=403)